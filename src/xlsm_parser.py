import os
import re
from openpyxl import load_workbook
from src.utils import columna_a_letra_excel

class LectorXLSM:
    def __init__(self, ruta_archivo):
        self.ruta_archivo = ruta_archivo
        self.libro = None
        self.codigo_establecimiento = None
        self.serie = None
        self.mes = None
        self._parsear_nombre_archivo()

    def _parsear_nombre_archivo(self):
        """
        Parsea el nombre del archivo para extraer el código de establecimiento, la serie y el mes.
        Formato esperado: [codigo_establecimiento][serie][mes].xlsm (ej: 116322A05.xlsm)
        """
        nombre_archivo = os.path.basename(self.ruta_archivo)
        coincidencia = re.match(r"(\d+)([A-Z]+)(\d{2})\.xlsm", nombre_archivo, re.IGNORECASE)
        if coincidencia:
            self.codigo_establecimiento = coincidencia.group(1)
            self.serie = coincidencia.group(2).upper()
            self.mes = coincidencia.group(3)
        else:
            raise ValueError(f"El nombre del archivo no coincide con el formato esperado: {nombre_archivo}")

    def cargar_libro(self):
        """Carga el libro de Excel."""
        try:
            self.libro = load_workbook(self.ruta_archivo, data_only=True, read_only=True)
            print(f"Archivo XLSM cargado: {self.ruta_archivo}")
        except Exception as e:
            print(f"Error al cargar el archivo XLSM {self.ruta_archivo}: {e}")
            self.libro = None

    def obtener_hoja(self, nombre_hoja):
        """Devuelve una hoja de cálculo específica por su nombre."""
        if self.libro is None:
            self.cargar_libro()
        if self.libro and nombre_hoja in self.libro.sheetnames:
            return self.libro[nombre_hoja]
        return None

    def obtener_valor_celda(self, nombre_hoja, fila, letra_col=None, indice_col=None):
        """
        Obtiene el valor de una celda específica.
        Se prefiere 'letra_col' (ej: 'A'), pero también se puede usar 'indice_col' (base 1).
        """
        hoja = self.obtener_hoja(nombre_hoja)
        if hoja:
            if letra_col:
                direccion_celda = f"{letra_col}{fila}"
            elif indice_col:
                letra_col = columna_a_letra_excel(indice_col)
                direccion_celda = f"{letra_col}{fila}"
            else:
                raise ValueError("Se debe proporcionar 'letra_col' o 'indice_col'.")
            
            celda = hoja[direccion_celda]
            return celda.value
        return None

    def obtener_valores_rango(self, nombre_hoja, celda_inicio, celda_fin):
        """
        Obtiene valores de un rango de celdas.
        'celda_inicio' y 'celda_fin' son referencias de celda estilo Excel (ej: 'A1', 'C11').
        Devuelve una lista de listas, representando filas y columnas.
        """
        hoja = self.obtener_hoja(nombre_hoja)
        if hoja:
            valores = []
            for fila in hoja[celda_inicio:celda_fin]:
                valores_fila = [celda.value for celda in fila]
                valores.append(valores_fila)
            return valores
        return None

    def cerrar_libro(self):
        """Cierra el libro si está abierto."""
        if self.libro:
            self.libro.close()
            self.libro = None
            print(f"Archivo XLSM cerrado: {self.ruta_archivo}")

# Ejemplo de uso (para fines de prueba)
if __name__ == "__main__":
    directorio_actual = os.path.dirname(__file__)
    ruta_archivo_xlsm = os.path.join(directorio_actual, '..', 'original', '116322A05.xlsm')

    try:
        lector = LectorXLSM(ruta_archivo_xlsm)
        print(f"Código de Establecimiento: {lector.codigo_establecimiento}")
        print(f"Serie: {lector.serie}")
        print(f"Mes: {lector.mes}")

        lector.cargar_libro()
        if lector.libro:
            valor_b2 = lector.obtener_valor_celda("NOMBRE", 2, letra_col='B')
            print(f"Valor de NOMBRE!B2: {valor_b2}")

            valores_c3_h3 = lector.obtener_valores_rango("NOMBRE", "C3", "H3")
            print(f"Valores de NOMBRE!C3:H3: {valores_c3_h3}")
        
        lector.cerrar_libro()

    except ValueError as ve:
        print(f"Error: {ve}")
    except Exception as e:
        print(f"Ocurrió un error inesperado: {e}")
